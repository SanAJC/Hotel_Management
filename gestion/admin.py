from django.contrib import admin
from .models import Huesped, Categoria, Producto , Habitacion, Venta
import openpyxl
from django.http import HttpResponse
from django import forms

def exportar_a_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Huéspedes"
    ws1.append(["ID", "Nombre", "Apellido", "Fecha de Check-in", "Fecha de Check-out"])  # Cabeceras
    
    for huesped in Huesped.objects.all():
        ws1.append([huesped.id, huesped.nombre, huesped.apellido, huesped.fecha_check_in, huesped.fecha_check_out])
    
    ws2 = wb.create_sheet(title="Inventario")
    ws2.append(["ID", "Nombre", "Categoría", "Cantidad", "Precio"])  # Cabeceras
    
    for producto in Producto.objects.all():
        ws2.append([producto.id, producto.nombre, str(producto.categoria), producto.cantidad, producto.precio])

    
    ws3 = wb.create_sheet(title="Ventas")
    ws3.append(["Fecha", "Producto", "Cantidad", "Total", "Huésped"])
    
    for venta in Venta.objects.all():
        ws3.append([
            venta.fecha,
            venta.producto.nombre,
            venta.cantidad,
            venta.total,
            str(venta.huesped) if venta.huesped else "Cliente General"
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="huespedes_inventario.xlsx"'
    
    wb.save(response)
    return response

# Describir la acción en la lista de acciones
exportar_a_excel.short_description = "Exportar huéspedes e inventario a Excel"

def check_out_huesped(modeladmin, request, queryset):
    for huesped in queryset:
        huesped.check_out()  # Llama a la función de check-out
    modeladmin.message_user(request, "Check-out realizado con éxito.")

check_out_huesped.short_description = "Realizar Check-out de huéspedes seleccionados"

@admin.register(Habitacion)
class HabitacionAdmin(admin.ModelAdmin):
    list_display = ('numero', 'disponible', 'precio', 'tiene_aire')
    list_filter = ('disponible',)
    search_fields = ('numero',)

    
class HuespedForm(forms.ModelForm):
    class Meta:
        model = Huesped
        exclude = ['fecha_check_out']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Filtrar solo las habitaciones disponibles
        self.fields['habitacion'].queryset = Habitacion.objects.filter(disponible=True)
        self.fields['habitacion'].label_from_instance = lambda obj: f'Habitación {obj.numero} - {"Con Aire" if obj.tiene_aire else "Sin Aire"} - Precio: ${obj.precio}'

@admin.register(Huesped)
class HuespedAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'apellido', 'cedula', 'habitacion', 'fecha_check_in', 'fecha_check_out', 'precio_pagado')
    list_filter = ('habitacion','nombre')
    search_fields = ('nombre', 'apellido','cedula')
    actions = [exportar_a_excel, check_out_huesped]
    form = HuespedForm  
    fields = ['nombre', 'apellido', 'cedula', 'habitacion', 'precio_pagado']
    readonly_fields = ['fecha_check_in']

    

@admin.register(Categoria)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ('nombre',)

@admin.register(Producto)
class ProductoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'cantidad', 'categoria')
    list_filter = ('categoria',)
    search_fields = ('nombre',)
    actions = [exportar_a_excel] 

@admin.register(Venta)
class VentaAdmin(admin.ModelAdmin):
    list_display = ('fecha', 'producto', 'cantidad', 'total', 'huesped')
    list_filter = ('fecha', 'producto', 'huesped')
    search_fields = ('producto__nombre', 'huesped__nombre', 'huesped__apellido')
    actions = [exportar_a_excel] 
    
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "producto":
            kwargs["queryset"] = Producto.objects.filter(cantidad__gt=0)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_readonly_fields(self, request, obj=None):
        if obj:  # Si estamos editando
            return ('fecha', 'total', 'producto', 'cantidad')
        return ('fecha', 'total')